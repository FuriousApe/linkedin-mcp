"""
Pre-processing script run once at Docker build time.
Reads lca.xlsx, filters to certified H-1B rows, writes one normalized
employer name per line to data/lca_employers.txt.
"""
import re
import sys
from pathlib import Path

import openpyxl

EXCEL = Path("data/lca.xlsx")
OUTPUT = Path("data/lca_employers.txt")

_SUFFIX = re.compile(
    r"[\s,]+(LLC|INC|INCORPORATED|CORP|CORPORATION|LTD|LP|LLP|PLLC|PC|CO|PTY|PLC)\b\.?$",
    re.IGNORECASE,
)


def normalize(name: str) -> str:
    name = name.upper().strip()
    name = _SUFFIX.sub("", name).strip().rstrip(",").strip()
    return name


def main() -> None:
    if not EXCEL.exists():
        print(f"ERROR: {EXCEL} not found", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {EXCEL} ...", flush=True)
    wb = openpyxl.load_workbook(str(EXCEL), read_only=True, data_only=True)
    ws = wb.active

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    name_col = headers.index("EMPLOYER_NAME")
    status_col = headers.index("CASE_STATUS")
    visa_col = headers.index("VISA_CLASS")

    employers: set[str] = set()
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if i % 100_000 == 0:
            print(f"  processed {i:,} rows, {len(employers):,} employers so far...", flush=True)
        status = str(row[status_col] or "")
        visa = str(row[visa_col] or "")
        if "Certified" in status and "H-1B" in visa:
            raw = row[name_col]
            if raw:
                employers.add(normalize(str(raw)))

    wb.close()

    OUTPUT.parent.mkdir(exist_ok=True)
    OUTPUT.write_text("\n".join(sorted(employers)), encoding="utf-8")
    print(f"Done — wrote {len(employers):,} certified H-1B employers to {OUTPUT}")


if __name__ == "__main__":
    main()
